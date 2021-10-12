import re
from os import startfile, path
from time import sleep
from datetime import date, datetime
from shutil import copyfile

import pandas as pd
import pyautogui as pya
from pyperclip import paste
from openpyxl import load_workbook

from constants import *


def get_datetime(date_or_time):

    if date_or_time == 'date':
        today = date.today().strftime("%d-%m-%Y")
        return today

    if date_or_time == 'time':
        now = datetime.now().strftime("%H:%M:%S")
        return now


def get_ticket_id_url():
    pya.hotkey('alt', 'd')
    sleep(.1)
    pya.hotkey('ctrl', 'c')
    sleep(.1)
    pya.press('tab')

    url = paste()

    if not url.startswith(LINK_STARTSwITH):
        return 'incorrect_link'

    ticket_id = re.sub('.*?([0-9]*)$', r'\1', url)

    return (ticket_id, url)


def get_excel_file():
    if not path.exists(f'./{EXCEL_NAME}'):

        df = pd.DataFrame({'Time': [],
                           'Ticket ID': [],
                           'Ticket link': []})
        writer = pd.ExcelWriter(EXCEL_NAME, engine='xlsxwriter')
        df.to_excel(writer, sheet_name=get_datetime('date'), index=False)
        writer.save()


def check_duplicate_ticket(workbook, id):
    ws = workbook[get_datetime('date')]
    last_row_id_cell = list(ws.iter_rows())[-1][-2]
    if last_row_id_cell.internal_value == id:
        return True


def edit_excel_sheet(workbook, data, ticket_id):

    if not CONSECUTIVE_DUPLICATES_ALLOWED:
        if check_duplicate_ticket(workbook, ticket_id):
            pya.alert(CUSTOM_ERRORS['duplicate'])
            return

    reader = pd.read_excel(EXCEL_NAME, get_datetime('date'))
    writer = pd.ExcelWriter(EXCEL_NAME, engine='openpyxl')
    writer.book = workbook
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    data.to_excel(writer, sheet_name=get_datetime('date'),
                  index=False, header=False, startrow=len(reader)+1)

    writer.close()


def add_new_excel_sheet(workbook, data):
    writer = pd.ExcelWriter(EXCEL_NAME, engine='openpyxl')
    writer.book = workbook
    data.to_excel(writer, sheet_name=get_datetime('date'), index=False)
    writer.book.active = len(writer.book.sheetnames) - 1

    writer.close()


def add_ticket():
    get_excel_file()

    clipboard_result = get_ticket_id_url()

    if type(clipboard_result) is str:

        if clipboard_result in CUSTOM_ERRORS.keys():
            pya.alert(CUSTOM_ERRORS[clipboard_result])
            return

    hyperlink = f'=HYPERLINK("{clipboard_result[1]}")'
    df = pd.DataFrame({'Time': [get_datetime('time')],
                       'Ticket ID': [clipboard_result[0]],
                       'Ticket link': [hyperlink]})

    book = load_workbook(EXCEL_NAME)

    if get_datetime('date') in book.sheetnames:
        edit_excel_sheet(book, df, clipboard_result[0])
    else:
        add_new_excel_sheet(book, df)


def get_today_ticket_count():
    try:
        reader = pd.read_excel(EXCEL_NAME, sheet_name=get_datetime('date'))
    except ValueError:
        pya.alert('No tickets for today yet.')
        return
    pya.alert(f'Tickets today: {reader.shape[0]}')


def open_sheet_copy():
    get_excel_file()

    try:
        copyfile(f'./{EXCEL_NAME}', './temporary.xlsx')
    except PermissionError:
        pya.alert('Close the previous file copy first.')
        return

    startfile('temporary.xlsx')
